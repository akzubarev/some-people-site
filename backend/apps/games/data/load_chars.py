import os
import sys
import traceback

import django
import openpyxl as openpyxl

sys.path[0] = '/app/'
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from apps.users.models import User
from apps.games.models import Game, Character, Tag, Faction


def load_game(game_alias: str):
    game = Game.objects.get(alias=game_alias)
    try:
        wb = openpyxl.load_workbook(
            filename=f"apps/games/data/{game_alias}-chars.xlsx",
            data_only=True
        )
        sheet = wb[wb.sheetnames[0]]
        row_num, column_num = 2, 1

        next_char = sheet.cell(row=row_num, column=column_num).value
        while next_char is not None:
            ch_al = next_char.split(", ")
            character_name = ch_al[0]
            alias = ch_al[1] if len(ch_al) > 1 else "-"
            master = sheet.cell(row=row_num, column=column_num + 1).value
            classs = sheet.cell(row=row_num, column=column_num + 2).value
            fraction = sheet.cell(row=row_num, column=column_num + 3).value
            description = sheet.cell(row=row_num, column=column_num + 4).value
            char, _ = Character.objects.get_or_create(
                name=character_name, alias=alias
            )
            master_user = User.objects.filter(id={
                "Леша": 2, "Настя": 3, "Аня": 4, "Катя": 5, "Лиза": 6
            }[master]).first()
            if master_user is not None:
                char.master = master_user
            if classs:
                tag, _ = Tag.objects.get_or_create(name=classs)
                char.tags.add(tag)
            faction, _ = Faction.objects.get_or_create(
                name=fraction, game=game
            )
            char.faction = faction
            char.description = description
            char.save()

            row_num += 1
            next_char = sheet.cell(row=row_num, column=column_num).value
    except Exception as e:
        traceback.print_exc()


def load_chars():
    for game_alias in ["frostpunk", "whales"]:
        load_game(game_alias=game_alias)


if __name__ == '__main__':
    load_chars()
